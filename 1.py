# 1/423536400
from math import fabs
import openpyxl
from ortools.linear_solver import pywraplp
import random
import os,sys
import time
from progress.bar import IncrementalBar

solver = pywraplp.Solver.CreateSolver('GLOP')

def solve(a, b, c1, c2):
    x = solver.NumVar(0, 50, 'x')
    y = solver.NumVar(0, 50, 'y')
    constr1 = solver.Constraint(0, b[0], 'ct')
    constr1.SetCoefficient(x, a[0])
    constr1.SetCoefficient(y, a[1])
    constr2 = solver.Constraint(0, b[1], 'ct')
    constr2.SetCoefficient(x, a[2])
    constr2.SetCoefficient(y, a[3])
    obje = solver.Objective()
    obje.SetCoefficient(x, c1)
    obje.SetCoefficient(y, c2)
    obje.SetMaximization()
    solver.Solve()
    return c1 * x.solution_value() + c2 * y.solution_value()


def setparam(a, b, cf, param_a, param_b, c1, c2):
    dia1 = ""
    dia2 = ""
    a[param_a] -= 1
    if cf > solve(a, b, c1, c2):
        a[param_a] += 1
        dia1 += "[" + str(a[param_a] - 8) + ".." + str(a[param_a]) + "]"

    else:
        a[param_a] += 2
        if cf > solve(a, b, c1, c2):
            a[param_a] -= 1
            dia1 += "[" + str(a[param_a]) + ".." + str(a[param_a] + 8) + "]"

        else:
            return ""

    b[param_b] -= 1
    if cf > solve(a, b, c1, c2):
        b[param_b] += 1
        dia2 += "[" + str(b[param_b] - 8) + ".." + str(b[param_b]) + "]"
    else:
        b[param_b] += 2
        if cf > solve(a, b, c1, c2):
            b[param_b] -= 1
            dia2 += "[" + str(b[param_b]) + ".." + str(b[param_b] + 8) + "]"
        else:
            return ""

    return "a: " + dia1 + " || b: " + dia2

try:
    os.remove("1.xlsx")
except:
    key = 0
key = 1
numel = 0
wb = openpyxl.Workbook()
while (True):
    try:
        print("Enter diapasone for c  : ")
        cd = int(input())
        print("Enter diapasone for a  : ")
        ad = int(input())
        print("Enter diapasone for b (for correct work b must be > 2) : ")
        bd = int(input())
        print("Enter num of expression  : ")
        numel = int(input())
        break
    except:
        print("Please, input correct data(integer value)")
bar = IncrementalBar('Progress', max = numel)
while (key != numel+1):
    constraints_num = 2

    x1 = solver.NumVar(0, 50, 'x1')  # variable creating
    x2 = solver.NumVar(0, 50, 'x2')

    b1 = random.randrange(2, bd, 1)
    ct1 = solver.Constraint(0, b1, 'ct')

    while (True):
        a1 = random.randrange(1, ad, 1)
        a2 = random.randrange(-ad, ad, 1)
        a3 = random.randrange(1, ad, 1)
        if a2 < 0:
            a4 = random.randrange(1, ad, 1)
        else:
            a4 = random.randrange(-ad, ad, 1)
        if a2 != 0 and a4 != 0:
            break
    ct1.SetCoefficient(x1, a1)
    ct1.SetCoefficient(x2, a2)
    b2 = random.randrange(2, bd, 1)
    ct2 = solver.Constraint(0, b2, 'ct')

    ct2.SetCoefficient(x1, a3)

    ct2.SetCoefficient(x2, a4)

    c1 = random.randrange(1, cd, 1)
    c2 = random.randrange(1, cd, 1)
    obj = solver.Objective()
    obj.SetCoefficient(x1, c1)
    obj.SetCoefficient(x2, c2)
    obj.SetMaximization()
    solver.Solve()
    xx1 = x1.solution_value()
    xx2 = x2.solution_value()

    if fabs(round(xx1) - xx1) < 0.0000000001 and fabs(round(xx2) - xx2) < 0.0000000001:
        cf = c1 * xx1 + c2 * xx2
        a = [a1, a2, a3, a4]
        b = [b1, b2]
        param_a = random.randrange(0, 4, 1)
        if param_a < 2:
            param_b = 1
        else:
            param_b = 0
        res = setparam(a, b, cf, param_a, param_b, c1, c2)

        try:
            point = [-(b[0] * a[3] - b[1] * a[1]) / (a[2] * a[1] - a[0] * a[3]),
                     (b[1] + (b[0] * a[3] - b[1] * a[1]) / (a[2] * a[1] - a[0] * a[3]) * a[2]) / a[3]
                     ]
        except:
            continue

        if res == "" or (
                fabs(round(point[0]) - point[0]) > 0.0000000001 or fabs(round(point[1]) - point[1]) > 0.0000000001 or (
                point[0] < 0 and point[1] < 0)):
            continue
        else:

            ws = wb.create_sheet(str(key))
            param = [a[param_a], b[param_b]]
            a[param_a] = "a"
            b[param_b] = "b"
            ws['A1'].value = str(c1) + "x1 + " + str(c2) + "x2 -> Max"
            ws['A2'].value = str(a[0]) + "x1 + " + str(a[1]) + "x2 <= " + str(b[0])
            ws['A3'].value = str(a[2]) + "x1 + " + str(a[3]) + "x2 <= " + str(b[1])
            ws['A4'].value = 'Solution:'
            ws['A5'].value = 'Objective value = ' + str(round(cf))
            ws['A6'].value = 'x1 =' + str(round(xx1))
            ws['A7'].value = 'x2 =' + str(round(xx2))
            ws['A8'].value = "a: " + str(param[0]) + " b: " + str(param[1])
            ws['A9'].value = res
            bar.next()
            time.sleep(0.1)
            key+=1
bar.finish()
sheet = wb['Sheet']
wb.remove(sheet)
wb.save("1.xlsx")
os.system("1.xlsx")
